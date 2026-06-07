import base64
import html as _html
import io
import json
import os
import re

import requests as req_lib
import streamlit as st
from datetime import date

try:
    from dotenv import load_dotenv

    load_dotenv()
except ImportError:
    pass

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
except ImportError:
    Workbook = None

BACKEND_URL = os.environ.get("BACKEND_URL", "http://127.0.0.1:8000")
APP_DIR = os.path.dirname(os.path.abspath(__file__))


# ─────────────────────────── helpers ───────────────────────────
def esc(value) -> str:
    if value is None:
        return ""
    return _html.escape(str(value))


def esc_multiline(value) -> str:
    return esc(value).replace("\n", "<br>")


def highlight(escaped_text: str, keyword: str) -> str:
    """Wrap case-insensitive keyword matches in <mark>. Operates on already-escaped HTML
    (the keyword is escaped too, so it lines up with the escaped text)."""
    if not keyword or not keyword.strip() or not escaped_text:
        return escaped_text
    esc_kw = esc(keyword.strip())
    if not esc_kw:
        return escaped_text
    return re.compile(re.escape(esc_kw), re.IGNORECASE).sub(
        lambda m: f'<mark class="kw-highlight">{m.group(0)}</mark>', escaped_text
    )


# Responsible team for each lifecycle phase (shown in the Detailed View findings).
PHASE_TO_TEAM = {
    "Coding Phase": "Engineering",
    "Test Phase": "System Test Team",
    "Requirement Phase": "Product Manager",
    "Design Review Phase": "Engineering",
    "Deployment Phase": "DevOps / Engineering",
    "Documentation Phase": "Service Readiness",
}


def help_dot(text: str) -> str:
    """A small, subtle round '?' that reveals `text` as a tooltip on hover."""
    return f'<span class="help-dot" title="{esc(text)}">?</span>'


def load_css():
    css_path = os.path.join(APP_DIR, "static", "styles.css")
    if os.path.exists(css_path):
        with open(css_path, "r", encoding="utf-8") as f:
            st.markdown(f"<style>{f.read()}</style>", unsafe_allow_html=True)


def b64_image(path: str) -> str:
    with open(path, "rb") as f:
        return base64.b64encode(f.read()).decode()


@st.cache_data(ttl=3600, show_spinner=False)
def load_discipline_product_map(nonce: int = 0) -> dict:
    """Fetch {discipline: [products]} map from backend.

    Returns a status dict {"ok": bool, "data": {...}, "error": str}. Successful
    results are cached for 1 hour; bumping `nonce` busts the cache so the reload
    button can force a fresh fetch. Failures are NOT cached.
    """
    try:
        resp = req_lib.get(f"{BACKEND_URL}/api/discipline-products", timeout=120)
        if resp.ok:
            return {"ok": True, "data": resp.json().get("discipline_products", {}), "error": ""}
        return {"ok": False, "data": {}, "error": f"Backend returned {resp.status_code}: {resp.text[:300]}"}
    except Exception as e:
        return {"ok": False, "data": {}, "error": str(e)}


def build_payload(mode, sprll_numbers, prompt_option, from_date=None, to_date=None, discipline=None, products=None, custom_jql=None, force_refresh=False):
    if mode == "Enter SPRLL Numbers manually":
        return {
            "sprll_numbers": sprll_numbers,
            "prompt_option": prompt_option,
            "force_refresh": force_refresh,
        }
    # Search by Date & Discipline: always carry discipline/products so the backend
    # can persist them on each gap (Database Insights) even when a custom JQL drives
    # the actual SPRLL resolution.
    payload = {
        "from_date": str(from_date) if from_date else None,
        "to_date": str(to_date) if to_date else None,
        "discipline": discipline,
        "products": products or [],
        "prompt_option": prompt_option,
        "force_refresh": force_refresh,
    }
    if custom_jql:
        payload["custom_jql"] = custom_jql
    return payload


def generate_missing_fields_excel(issues):
    """Generate an Excel report of missing fields per SPRLL issue."""
    if Workbook is None:
        return None

    wb = Workbook()
    ws = wb.active
    ws.title = "Missing Fields Report"

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    error_fill = PatternFill(start_color="FEE2E2", end_color="FEE2E2", fill_type="solid")
    error_font = Font(color="9F1239")
    wrap_alignment = Alignment(wrap_text=True, vertical="top")

    headers = ["SPRLL Number", "Summary", "Status", "Discipline", "Missing Fields", "Remarks"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = wrap_alignment

    row = 2
    for issue in issues:
        sprll_num = issue.get("sprllNumber", issue.get("key", ""))
        has_error = issue.get("error")

        ws.cell(row=row, column=1, value=sprll_num).alignment = wrap_alignment

        if has_error:
            ws.cell(row=row, column=2, value="—").alignment = wrap_alignment
            ws.cell(row=row, column=3, value="—").alignment = wrap_alignment
            ws.cell(row=row, column=4, value="—").alignment = wrap_alignment
            ws.cell(row=row, column=5, value="—").alignment = wrap_alignment
            remark_cell = ws.cell(row=row, column=6, value=f"Not a proper SPRLL number or fetch error: {has_error}")
            remark_cell.alignment = wrap_alignment
            remark_cell.fill = error_fill
            remark_cell.font = error_font
        else:
            summary = issue.get("summary", "—")
            status = issue.get("status", "—")
            discipline = issue.get("discipline", issue.get("customfield_12801", "—"))
            missing = issue.get("missingFields", issue.get("missing_fields", []))
            missing_str = ", ".join(missing) if missing else "None"

            ws.cell(row=row, column=2, value=summary).alignment = wrap_alignment
            ws.cell(row=row, column=3, value=status).alignment = wrap_alignment
            ws.cell(row=row, column=4, value=discipline).alignment = wrap_alignment
            ws.cell(row=row, column=5, value=missing_str).alignment = wrap_alignment

            if missing:
                remark = f"{len(missing)} field(s) missing"
                remark_cell = ws.cell(row=row, column=6, value=remark)
                remark_cell.alignment = wrap_alignment
                remark_cell.fill = error_fill
                remark_cell.font = error_font
            else:
                ws.cell(row=row, column=6, value="All required fields present").alignment = wrap_alignment

        row += 1

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 40
    ws.column_dimensions["C"].width = 14
    ws.column_dimensions["D"].width = 30
    ws.column_dimensions["E"].width = 45
    ws.column_dimensions["F"].width = 45
    ws.freeze_panes = "A2"

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


def render_database_insights():
    """Render the recurring-gaps Database Insights view (slicing + cluster cards).

    Reused on the landing screen (before any analysis is run) and inside a collapsed
    expander below the result tabs once an analysis exists.
    """
    st.markdown(
        '<div class="section-sub">🔁 Most-repeated Action items across <strong>all saved '
        'analyses</strong>. Switch the slice to explore recurring themes.</div>',
        unsafe_allow_html=True,
    )
    try:
        dim_resp = req_lib.get(f"{BACKEND_URL}/api/gap-dimensions", timeout=60)
        dims = dim_resp.json() if dim_resp.ok else {}
    except Exception:
        dims = {}

    dim_label_to_key = {
        "By Discipline": "discipline",
        "By Product": "product",
    }
    slice_col, value_col, phase_col = st.columns([1.3, 1, 1])
    with slice_col:
        dim_label = st.radio(
            "Slice by",
            list(dim_label_to_key.keys()),
            horizontal=True,
            key="db_dim",
        )
    dim_key = dim_label_to_key[dim_label]
    value_options = {
        "discipline": dims.get("disciplines", []),
        "product": dims.get("products", []),
    }[dim_key]
    with value_col:
        selected_value = st.selectbox(
            "Filter value", ["All"] + value_options, key=f"db_value_{dim_key}"
        )
    with phase_col:
        phase_filter = st.selectbox(
            "Lifecycle phase",
            ["All"] + dims.get("lifecycle_phases", []),
            key="db_phase_filter",
        )

    keyword = st.text_input(
        "🔍 Keyword search",
        placeholder="Find a word anywhere in saved insights & SPRLLs…",
        key="db_search_keyword",
    ).strip()

    issues = []
    try:
        with st.spinner("Loading database insights…"):
            ins_resp = req_lib.post(
                f"{BACKEND_URL}/api/gap-insights",
                json={
                    "dimension": dim_key,
                    "value": selected_value,
                    "lifecycle_phase": phase_filter,
                    "min_cluster_size": 1,
                    "top_n": 25,
                    "keyword": keyword,
                },
                timeout=120,
            )
        if ins_resp.ok:
            payload = ins_resp.json()
            clusters = payload.get("clusters", [])
            issues = payload.get("issues", [])
        else:
            clusters = []
            st.error(f"Could not load insights: {ins_resp.status_code} {ins_resp.text[:200]}")
    except Exception as e:
        clusters = []
        st.error(f"Could not load insights: {e}")

    if not clusters and not issues:
        if keyword:
            msg = f"No saved insights or SPRLLs match “{esc(keyword)}”."
        else:
            msg = ("No saved action items yet for this slice. "
                   "Run analyses to populate the database.")
        st.markdown(f'<div class="empty-state">{msg}</div>', unsafe_allow_html=True)
        return

    if clusters:
        st.markdown(
            f'<div class="section-sub">🏆 Top {len(clusters)} recurring action item(s)</div>',
            unsafe_allow_html=True,
        )
    for rank, cl in enumerate(clusters, 1):
        rep = cl.get("representative", {})
        r_phase = rep.get("lifecycle_phase", "")
        r_score = rep.get("validation_score", "")
        r_title = rep.get("title", "Untitled")
        r_action = (
            rep.get("improved_recommendation")
            or rep.get("recommended_fix")
            or rep.get("description")
            or ""
        )
        r_disc = ", ".join(cl.get("disciplines", []))
        r_prods = ", ".join(cl.get("products", []))
        r_sprlls = cl.get("source_sprll_keys", [])

        d_badges = []
        if r_phase:
            d_badges.append(f'<span class="badge badge-purple">⏱ {esc(r_phase)}</span>')
        if r_score not in ("", None):
            d_badges.append(f'<span class="badge badge-neutral">⚖ {esc(r_score)}/5</span>')

        d_card = ['<div class="gap-card">']
        d_card.append('<div class="gap-card-header">')
        d_card.append(f'<div class="gap-number">{rank}</div>')
        d_card.append('<div class="gap-title-wrap">')
        if d_badges:
            d_card.append(f'<div class="gap-meta">{"".join(d_badges)}</div>')
        d_card.append(f'<div class="gap-title">{highlight(esc(r_title), keyword)}</div>')
        d_card.append('</div></div>')
        if r_action:
            d_card.append(
                '<div class="gap-section">'
                '<div class="gap-section-label">✅ Recommended QA Action</div>'
                f'<div class="gap-section-body">{highlight(esc_multiline(r_action), keyword)}</div>'
                '</div>'
            )
        if r_disc:
            d_card.append(
                '<div class="gap-section">'
                '<div class="gap-section-label">🎯 Disciplines</div>'
                f'<div class="gap-section-body">{highlight(esc(r_disc), keyword)}</div>'
                '</div>'
            )
        if r_prods:
            d_card.append(
                '<div class="gap-section">'
                '<div class="gap-section-label">📦 Products</div>'
                f'<div class="gap-section-body">{highlight(esc(r_prods), keyword)}</div>'
                '</div>'
            )
        if r_sprlls:
            d_chips = "".join(
                f'<span class="sprll-chip">{esc(k)}</span>' for k in r_sprlls
            )
            d_card.append(
                '<div class="gap-section">'
                f'<div class="gap-section-label">🔗 Across SPRLLs ({len(r_sprlls)})</div>'
                f'<div class="sprll-chips">{d_chips}</div>'
                '</div>'
            )
        d_card.append('</div>')
        st.markdown("".join(d_card), unsafe_allow_html=True)

    if keyword and issues:
        st.markdown(
            f'<div class="section-sub">🔎 {len(issues)} matching SPRLL(s)</div>',
            unsafe_allow_html=True,
        )
        for iss in issues:
            i_key = iss.get("key", "")
            i_status = iss.get("status", "")
            i_summary = iss.get("summary", "")
            i_desc = (iss.get("description") or "")[:400]

            i_meta = [f'<span class="sprll-chip">{esc(i_key)}</span>']
            if i_status:
                i_meta.append(f'<span class="badge badge-neutral">{esc(i_status)}</span>')

            i_card = ['<div class="gap-card">']
            i_card.append('<div class="gap-card-header">')
            i_card.append('<div class="gap-title-wrap">')
            i_card.append(f'<div class="gap-meta">{"".join(i_meta)}</div>')
            i_card.append(
                f'<div class="gap-title">{highlight(esc(i_summary), keyword)}</div>'
            )
            i_card.append('</div></div>')
            if i_desc:
                i_card.append(
                    '<div class="gap-section">'
                    '<div class="gap-section-label">📝 Description</div>'
                    f'<div class="gap-section-body">{highlight(esc_multiline(i_desc), keyword)}</div>'
                    '</div>'
                )
            i_card.append('</div>')
            st.markdown("".join(i_card), unsafe_allow_html=True)


# ─────────────────────────── page setup ───────────────────────────
st.set_page_config(
    page_title="SPRLL Process Gap Analyzer",
    page_icon="🚀",
    layout="wide",
    initial_sidebar_state="collapsed",
)
load_css()


# ─────────────────────────── hero header ───────────────────────────
logo_path = os.path.join(APP_DIR, "static", "zebra_logo.jpeg")
logo_html = ""
if os.path.exists(logo_path):
    try:
        logo_b64 = b64_image(logo_path)
        logo_html = (
            f'<div class="hero-logo">'
            f'<img src="data:image/jpeg;base64,{logo_b64}" '
            f'style="height:46px;display:block;" alt="Zebra"/>'
            f'</div>'
        )
    except Exception:
        logo_html = ""

st.markdown(
    f"""
    <div class="hero-header">
      <div class="hero-content">
        <div>
          <div class="hero-badge"><span class="dot"></span> Platform for Proactive Quality Assurance</div>
          <h1 class="hero-title">SPRLL Process Gap Analyzer</h1>
          <div class="hero-subtitle">
            AI-powered analysis of release-blocking signals across SPRLL issues —
            uncovering systemic gaps, validating findings, and driving smarter releases.
          </div>
        </div>
        {logo_html}
      </div>
    </div>
    """,
    unsafe_allow_html=True,
)


# ─────────────── database insights (collapsible, above parameters) ───────────────
with st.expander("🗂️ Database Insights", expanded=False):
    render_database_insights()


# ─────────────────────────── analysis parameters ───────────────────────────
st.markdown('<div class="section-header">Analysis Parameters</div>', unsafe_allow_html=True)

prompt_option = 1  # fixed (Senior Analyst)

mode = st.radio(
    "Input mode",
    ["Enter SPRLL Numbers manually", "Search by Date & Discipline"],
    horizontal=True,
    label_visibility="collapsed",
)

sprll_numbers = []
from_date = None
to_date = None
discipline = None
selected_products = None
jql = None

if mode == "Enter SPRLL Numbers manually":
    sprll_raw = st.text_input(
        "SPRLL Number(s)",
        placeholder="e.g.  4212, 3974, 4112",
        help="Separate multiple numbers with commas or spaces. The 'SPRLL-' prefix is added automatically.",
    )
    if sprll_raw:
        parts = re.split(r"[,\s]+", sprll_raw.strip())
        for p in parts:
            p = p.strip()
            if not p:
                continue
            num = re.sub(r"(?i)^sprll-?", "", p)
            if num:
                sprll_numbers.append(f"SPRLL-{num}")

    if sprll_numbers:
        chips_html = " ".join(f'<span class="sprll-chip">{esc(n)}</span>' for n in sprll_numbers)
        st.markdown(
            f'<div style="margin-top:0.6rem;"><span class="field-label">Will analyze</span>'
            f'<div class="sprll-chips" style="margin-top:0.35rem;">{chips_html}</div></div>',
            unsafe_allow_html=True,
        )
else:
    disc_nonce = st.session_state.get("disc_nonce", 0)
    with st.spinner("Loading disciplines & products from Jira…"):
        disc_result = load_discipline_product_map(disc_nonce)

    if not disc_result.get("ok"):
        st.error(
            f"Could not load disciplines & products from Jira: {disc_result.get('error', 'unknown error')}"
        )
        if st.button("🔄  Reload disciplines & products"):
            st.session_state["disc_nonce"] = disc_nonce + 1
            st.rerun()
        st.stop()

    disc_prod_map = disc_result.get("data", {})
    available_disciplines = sorted(disc_prod_map.keys()) if disc_prod_map else []

    c1, c2, c3 = st.columns(3)
    with c1:
        from_date = st.date_input("From Date", value=date(date.today().year, 1, 1))
    with c2:
        to_date = st.date_input("To Date", value=date.today())
    with c3:
        discipline = st.selectbox(
            "Discipline",
            options=available_disciplines,
            help="Disciplines are loaded live from Jira.",
        )

    available_products = disc_prod_map.get(discipline, []) if discipline else []

    selected_products = st.multiselect(
        "Product (optional)",
        options=available_products,
        placeholder="Select one or more products…" if available_products else "No products for this discipline",
        help="Products are filtered by the selected discipline, loaded live from Jira.",
    )

    disc_clause = f'type in (Process, "Lesson Learned") AND Discipline = "{discipline}"'
    date_filter = f'created >= "{from_date}" AND created <= "{to_date}"'
    status_filter = "status in (Resolved, Closed)"

    if selected_products:
        prod_list = ", ".join(f'"{p}"' for p in selected_products)
        prod_clause = f'type in (Process, "Lesson Learned") AND "Product Selection" in ({prod_list})'
        default_jql = (
            f"project = SPRLL AND "
            f"({disc_clause} OR {prod_clause}) AND "
            f"{status_filter} AND {date_filter}"
        )
    else:
        default_jql = (
            f"project = SPRLL AND {disc_clause} AND "
            f"{status_filter} AND {date_filter}"
        )

    with st.expander("JQL Query Preview", expanded=True):
        jql = st.text_area("Edit the JQL query if needed:", value=default_jql, height=100)

st.markdown('<div style="margin-top:1rem;"></div>', unsafe_allow_html=True)
btn_col1, btn_col2, help_col, _ = st.columns([1, 1, 0.4, 2.6])
with btn_col1:
    analyze_clicked = st.button("🚀  Run Analysis", type="primary", use_container_width=True)
with btn_col2:
    force_clicked = st.button("🔁 Re-Fetch", use_container_width=True)
with help_col:
    st.markdown(
        f'<div class="help-inline-btn">{help_dot("Re-Fetch ignores any saved result for this query and regenerates the analysis from scratch with the AI model — use it when the SPRLL data changed or you want a fresh run instead of the cached result.")}</div>',
        unsafe_allow_html=True,
    )


# ─────────────────────────── analysis output ───────────────────────────
if analyze_clicked or force_clicked:
    payload = build_payload(
        mode=mode,
        sprll_numbers=sprll_numbers,
        prompt_option=prompt_option,
        from_date=from_date,
        to_date=to_date,
        discipline=discipline,
        products=selected_products,
        custom_jql=jql,
        force_refresh=force_clicked,
    )

    try:
        spinner_msg = (
            "Generating analysis with Vertex AI · This may take a moment…"
            if force_clicked
            else "Analyzing SPRLL issues with Vertex AI · This may take a moment…"
        )
        with st.spinner(spinner_msg):
            resp = req_lib.post(f"{BACKEND_URL}/api/analyze", json=payload, timeout=180)

        if not resp.ok:
            st.error(f"Backend returned {resp.status_code}: {resp.text[:500]}")
        else:
            st.session_state["analysis_result"] = resp.json()
            st.session_state["analysis_meta"] = {
                "mode": mode,
                "from_date": str(from_date) if from_date else "",
                "to_date": str(to_date) if to_date else "",
                "discipline": discipline or "",
            }
    except Exception as e:
        st.error(f"Analysis failed: {e}")


# ─────────── render output (persisted in session so panels stay interactive) ───────────
result = st.session_state.get("analysis_result")
if result:
    meta = st.session_state.get("analysis_meta", {})
    try:
        issues = result.get("issues", [])
        process_gaps = result.get("process_gaps", [])
        resolved_sprll_numbers = result.get("sprll_numbers", [])
        st.session_state["last_issues"] = issues

        if result.get("cached"):
            st.caption("⚡ Served from a saved analysis — use **Re-Fetch** to regenerate.")

        # ───── Scope bar ─────
        if meta.get("mode") == "Search by Date & Discipline":
            scope_html = (
                f'<div class="scope-bar">'
                f'<span class="scope-item">📅 <strong>{esc(meta.get("from_date"))}</strong> → <strong>{esc(meta.get("to_date"))}</strong></span>'
                f'<span class="scope-divider">•</span>'
                f'<span class="scope-item">🎯 Discipline: <strong>{esc(meta.get("discipline"))}</strong></span>'
                f'<span class="scope-divider">•</span>'
                f'<span class="scope-item"><strong>{len(resolved_sprll_numbers)}</strong> issue(s) in scope</span>'
                f'</div>'
            )
        else:
            scope_html = (
                f'<div class="scope-bar">'
                f'<span class="scope-item">🎯 <strong>{len(resolved_sprll_numbers)}</strong> SPRLL issue(s) analyzed manually</span>'
                f'</div>'
            )
        st.markdown(scope_html, unsafe_allow_html=True)

        # ───── KPI cards ─────
        total_issues = len(issues)
        total_gaps = len(process_gaps)

        compliant = sum(
            1 for i in issues
            if not i.get("error") and not (i.get("missingFields") or i.get("missing_fields") or [])
        )
        analyzed_non_error = sum(1 for i in issues if not i.get("error"))
        compliance_pct = int(round((compliant / analyzed_non_error) * 100)) if analyzed_non_error else 0

        validations = [g.get("validation") for g in process_gaps if g.get("validation")]
        valid_count = sum(1 for v in validations if (v.get("validation_result") or "").lower() == "valid")
        validated_pct = int(round((valid_count / len(validations)) * 100)) if validations else None

        kpi_html = '<div class="kpi-row">'
        kpi_html += (
            f'<div class="kpi-card primary">'
            f'<div class="kpi-label">Issues Analyzed</div>'
            f'<div class="kpi-value">{total_issues}</div>'
            f'<div class="kpi-sub">SPRLL items in this run</div>'
            f'</div>'
        )
        kpi_html += (
            f'<div class="kpi-card info">'
            f'<div class="kpi-label">Action Items</div>'
            f'<div class="kpi-value">{total_gaps}</div>'
            f'<div class="kpi-sub">QA actions for upcoming releases</div>'
            f'</div>'
        )
        kpi_html += (
            f'<div class="kpi-card success">'
            f'<div class="kpi-label">Field Compliance</div>'
            f'<div class="kpi-value">{compliance_pct}%</div>'
            f'<div class="kpi-sub">{compliant} of {analyzed_non_error} fully populated</div>'
            f'</div>'
        )
        if validated_pct is not None:
            kpi_html += (
                f'<div class="kpi-card warning">'
                f'<div class="kpi-label">Judge Validated</div>'
                f'<div class="kpi-value">{validated_pct}%</div>'
                f'<div class="kpi-sub">{valid_count} of {len(validations)} rated <em>Valid</em></div>'
                f'</div>'
            )
        kpi_html += '</div>'
        st.markdown(kpi_html, unsafe_allow_html=True)

        # ───── 2-panel output ─────
        tab_quick, tab_full = st.tabs([
            "  ⚡ Quick View  ",
            "  📋 Detailed View  ",
        ])

        # ───── Panel 1: Quick View ─────
        with tab_quick:
            if not process_gaps:
                st.markdown(
                    '<div class="empty-state">No action items were generated for this scope.</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    "<div class=\"section-sub\">⚡Action items for upcoming "
                    "releases, each shown with the LLM judge's improved recommendation.</div>",
                    unsafe_allow_html=True,
                )
                for q_idx, gap in enumerate(process_gaps, 1):
                    q_num = gap.get("number", q_idx)
                    q_title = gap.get("title", "Untitled")
                    q_phase = gap.get("lifecycle_phase") or gap.get("phase", "")
                    q_val = gap.get("validation") or {}
                    q_result = q_val.get("validation_result", "")
                    q_score = q_val.get("validation_score", "")
                    q_action = q_val.get("improved_recommendation") or gap.get("recommended_fix", "")

                    q_badges = []
                    if q_phase:
                        q_badges.append(f'<span class="badge badge-purple">⏱ {esc(q_phase)}</span>')
                    if q_result:
                        q_badge_class = {
                            "Valid": "badge-success",
                            "Partially Valid": "badge-warning",
                            "Invalid": "badge-danger",
                        }.get(q_result, "badge-neutral")
                        q_score_txt = f" · {esc(q_score)}/5" if q_score != "" else ""
                        q_badges.append(
                            f'<span class="badge {q_badge_class}">⚖ {esc(q_result)}{q_score_txt}</span>'
                        )

                    q_card = ['<div class="gap-card">']
                    q_card.append('<div class="gap-card-header">')
                    q_card.append(f'<div class="gap-number">{esc(q_num)}</div>')
                    q_card.append('<div class="gap-title-wrap">')
                    if q_badges:
                        q_card.append(f'<div class="gap-meta">{"".join(q_badges)}</div>')
                    q_card.append(f'<div class="gap-title">{esc(q_title)}</div>')
                    q_card.append('</div></div>')
                    if q_action:
                        q_card.append(
                            '<div class="gap-section">'
                            '<div class="gap-section-label">✅ Recommendation </div>'
                            f'<div class="gap-section-body">{esc_multiline(q_action)}</div>'
                            '</div>'
                        )
                    q_card.append('</div>')
                    st.markdown("".join(q_card), unsafe_allow_html=True)

        # ───── Panel 2: Detailed View (sub-tabs) ─────
        with tab_full:
            st.markdown(
                '<div class="dv-help">' + help_dot(
                    "Findings are process gaps synthesized from the combined SPRLL evidence, "
                    "each scored 0–5 by an LLM judge (the ⚖ badge) and tagged with the "
                    "responsible team. Analyzed Issues lists every SPRLL in this run with its "
                    "field-compliance status and comments, plus the missing-fields Excel report."
                ) + '</div>',
                unsafe_allow_html=True,
            )
            tab_gaps, tab_issues = st.tabs([
                f"  Findings ({total_gaps})  ",
                f"  Analyzed Issues ({total_issues})  ",
            ])

        # ───── Gaps tab ─────
        with tab_gaps:
            if not process_gaps:
                st.markdown(
                    '<div class="empty-state">No gaps were generated for this scope.</div>',
                    unsafe_allow_html=True,
                )
            else:
                st.markdown(
                    '<div class="section-sub">Findings synthesized from combined SPRLL evidence, '
                    'each independently scored by an LLM judge.</div>',
                    unsafe_allow_html=True,
                )

                for idx, gap in enumerate(process_gaps, 1):
                    num = gap.get("number", idx)
                    title = gap.get("title", "Untitled")
                    desc = gap.get("description") or gap.get("gap_description", "")
                    evidence = gap.get("evidence", "")
                    fix = gap.get("recommended_fix", "")
                    lifecycle_phase = gap.get("lifecycle_phase") or gap.get("phase", "")
                    related = gap.get("related_sprll") or []
                    validation = gap.get("validation") or {}

                    meta_badges = []
                    if lifecycle_phase:
                        meta_badges.append(
                            f'<span class="badge badge-purple">⏱ {esc(lifecycle_phase)}</span>'
                        )
                        team = PHASE_TO_TEAM.get(lifecycle_phase)
                        if team:
                            meta_badges.append(
                                f'<span class="badge badge-team">👥 {esc(team)}</span>'
                            )
                    v_result = validation.get("validation_result", "")
                    v_score = validation.get("validation_score", "")
                    if v_result:
                        badge_class = {
                            "Valid": "badge-success",
                            "Partially Valid": "badge-warning",
                            "Invalid": "badge-danger",
                        }.get(v_result, "badge-neutral")
                        score_txt = f" · {esc(v_score)}/5" if v_score != "" else ""
                        meta_badges.append(
                            f'<span class="badge {badge_class}">⚖ {esc(v_result)}{score_txt}</span>'
                        )

                    keys = []
                    for e in related:
                        if isinstance(e, dict):
                            k = e.get("key", "")
                        else:
                            k = str(e)
                        if k:
                            keys.append(k)
                    chips_html = "".join(f'<span class="sprll-chip">{esc(k)}</span>' for k in keys)

                    card = ['<div class="gap-card">']
                    card.append('<div class="gap-card-header">')
                    card.append(f'<div class="gap-number">{esc(num)}</div>')
                    card.append('<div class="gap-title-wrap">')
                    if meta_badges:
                        card.append(f'<div class="gap-meta">{"".join(meta_badges)}</div>')
                    card.append(f'<div class="gap-title">{esc(title)}</div>')
                    card.append('</div></div>')  # /title-wrap, /header

                    if desc:
                        card.append(f'<div class="gap-desc">{esc_multiline(desc)}</div>')

                    if evidence:
                        card.append(
                            '<div class="gap-section">'
                            '<div class="gap-section-label">📌 Evidence</div>'
                            f'<div class="gap-section-body italic">{esc_multiline(evidence)}</div>'
                            '</div>'
                        )
                    if fix:
                        card.append(
                            '<div class="gap-section">'
                            '<div class="gap-section-label">✅ Recommended Fix</div>'
                            f'<div class="gap-section-body">{esc_multiline(fix)}</div>'
                            '</div>'
                        )
                    if chips_html:
                        card.append(
                            '<div class="gap-section">'
                            f'<div class="gap-section-label">🔗 Related SPRLLs ({len(keys)})</div>'
                            f'<div class="sprll-chips">{chips_html}</div>'
                            '</div>'
                        )

                    card.append('</div>')  # /gap-card
                    st.markdown("".join(card), unsafe_allow_html=True)

                    # Validation drill-down (as Streamlit expander placed right after)
                    if validation:
                        v_color = {
                            "Valid": "🟢",
                            "Partially Valid": "🟡",
                            "Invalid": "🔴",
                        }.get(v_result, "⚪")
                        label = f"{v_color}  LLM Judge Details — {v_result or 'N/A'}"
                        if v_score != "":
                            label += f"  (Score {v_score}/5)"
                        with st.expander(label):
                            v_grid = ['<div class="field-grid">']
                            persona = validation.get("assigned_persona")
                            conf = validation.get("confidence")
                            if persona:
                                v_grid.append(
                                    f'<div class="field-item"><div class="field-label">Persona</div>'
                                    f'<div class="field-value">{esc(persona)}</div></div>'
                                )
                            if conf:
                                v_grid.append(
                                    f'<div class="field-item"><div class="field-label">Confidence</div>'
                                    f'<div class="field-value mono">{esc(conf)}</div></div>'
                                )
                            v_grid.append('</div>')
                            st.markdown("".join(v_grid), unsafe_allow_html=True)

                            if validation.get("reason"):
                                st.markdown(
                                    '<div class="gap-section-label">Reasoning</div>',
                                    unsafe_allow_html=True,
                                )
                                st.markdown(validation["reason"])
                            v_issues = validation.get("identified_issues") or []
                            if v_issues:
                                st.markdown(
                                    '<div class="gap-section-label" style="margin-top:0.8rem;">Identified Issues</div>',
                                    unsafe_allow_html=True,
                                )
                                for vi in v_issues:
                                    st.markdown(f"- {vi}")
                            improved = validation.get("improved_recommendation", "")
                            if improved:
                                st.markdown(
                                    '<div class="gap-section-label" style="margin-top:0.8rem;">Improved Recommendation</div>',
                                    unsafe_allow_html=True,
                                )
                                st.markdown(improved)

        # ───── Issues tab ─────
        with tab_issues:
            if Workbook is not None:
                excel_buffer = generate_missing_fields_excel(issues)
                if excel_buffer:
                    col_dl, col_dl_help, _ = st.columns([1, 0.4, 2.6])
                    with col_dl:
                        st.download_button(
                            label="📥  Download Missing Fields Report",
                            data=excel_buffer,
                            file_name="sprll_missing_fields_report.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            use_container_width=True,
                        )
                    with col_dl_help:
                        st.markdown(
                            f'<div class="help-inline-btn">{help_dot("Exports every analyzed SPRLL and the required fields it is missing to an Excel workbook, so gaps can be tracked and assigned offline.")}</div>',
                            unsafe_allow_html=True,
                        )
            else:
                st.warning("Install `openpyxl` to enable Excel download: `pip install openpyxl`")

            if not issues:
                st.markdown(
                    '<div class="empty-state">No issues to display.</div>',
                    unsafe_allow_html=True,
                )

            for issue in issues:
                sprll_num = issue.get("sprllNumber", issue.get("key", "SPRLL"))
                summary = issue.get("summary", "")
                missing = issue.get("missingFields", issue.get("missing_fields", []))
                has_error = issue.get("error")

                if has_error:
                    status_emoji = "❌"
                elif missing:
                    status_emoji = "⚠️"
                else:
                    status_emoji = "✅"

                label = f"{status_emoji}  {sprll_num}  —  {summary[:120]}"

                with st.expander(label, expanded=False):
                    if has_error:
                        st.error(issue["error"])
                        continue

                    status = issue.get("status", "—")
                    discipline_val = issue.get("discipline", issue.get("customfield_12801", "—"))
                    assignee = issue.get("assignee_name", "—")
                    comment_count = issue.get(
                        "matchedCommentCount", issue.get("comment_count", 0)
                    )

                    grid = ['<div class="field-grid">']
                    grid.append(
                        f'<div class="field-item"><div class="field-label">Status</div>'
                        f'<div class="field-value">{esc(status)}</div></div>'
                    )
                    grid.append(
                        f'<div class="field-item"><div class="field-label">Discipline</div>'
                        f'<div class="field-value">{esc(discipline_val)}</div></div>'
                    )
                    grid.append(
                        f'<div class="field-item"><div class="field-label">Assignee</div>'
                        f'<div class="field-value">{esc(assignee)}</div></div>'
                    )
                    grid.append(
                        f'<div class="field-item"><div class="field-label">Assignee Comments</div>'
                        f'<div class="field-value">{esc(comment_count)}</div></div>'
                    )
                    grid.append('</div>')
                    st.markdown("".join(grid), unsafe_allow_html=True)

                    # compliance badge
                    if missing:
                        bad_html = (
                            f'<div style="margin-top:0.25rem;">'
                            f'<span class="badge badge-warning">⚠ {len(missing)} field(s) missing</span>'
                            f'<div style="margin-top:0.4rem;color:#92400e;font-size:0.88rem;">'
                            f'{esc(", ".join(missing))}</div></div>'
                        )
                    else:
                        bad_html = (
                            '<div style="margin-top:0.25rem;">'
                            '<span class="badge badge-success">✓ All required fields present</span>'
                            '</div>'
                        )
                    st.markdown(bad_html, unsafe_allow_html=True)

                    source = issue.get("source", "")
                    if source:
                        st.markdown(
                            f'<div class="tiny muted" style="margin-top:0.6rem;">📦 Source: '
                            f'<code>{esc(source)}</code></div>',
                            unsafe_allow_html=True,
                        )

                    desc = issue.get("description", "")
                    if desc:
                        st.markdown(
                            '<div class="gap-section-label" style="margin-top:1rem;">Description</div>',
                            unsafe_allow_html=True,
                        )
                        st.markdown(desc[:800] + ("…" if len(desc) > 800 else ""))

                    assignee_comments = issue.get("assignee_comments", [])
                    if assignee_comments:
                        st.markdown(
                            f'<div class="gap-section-label" style="margin-top:1rem;">'
                            f'Assignee Comments ({len(assignee_comments)})</div>',
                            unsafe_allow_html=True,
                        )
                        for i, comment_text in enumerate(assignee_comments, 1):
                            st.markdown(
                                f"**Comment {i}:**\n\n"
                                + comment_text[:600]
                                + ("…" if len(comment_text) > 600 else "")
                            )

                    gen_summary = issue.get(
                        "assigneeCommentSummary",
                        issue.get("generated_summary", ""),
                    )
                    if gen_summary:
                        st.markdown(
                            '<div class="gap-section-label" style="margin-top:1rem;">🤖 AI Summary</div>',
                            unsafe_allow_html=True,
                        )
                        try:
                            summary_json = json.loads(gen_summary)
                            if isinstance(summary_json, dict):
                                if "summary" in summary_json:
                                    st.markdown(f"**Summary:** {summary_json['summary']}")
                                if "key_points" in summary_json:
                                    st.markdown("**Key Points:**")
                                    for kp in summary_json["key_points"]:
                                        st.markdown(f"- {kp}")
                                if "key_actions_or_decisions" in summary_json:
                                    st.markdown("**Key Actions / Decisions:**")
                                    for ka in summary_json["key_actions_or_decisions"]:
                                        st.markdown(f"- {ka}")
                                if "evidence" in summary_json:
                                    st.markdown("**Evidence:**")
                                    if isinstance(summary_json["evidence"], list):
                                        for ev in summary_json["evidence"]:
                                            st.markdown(f"- _{ev}_")
                                    else:
                                        st.markdown(f"_{summary_json['evidence']}_")
                                if "confidence" in summary_json:
                                    st.markdown(f"**Confidence:** `{summary_json['confidence']}`")
                                if "limitations" in summary_json and summary_json["limitations"]:
                                    st.markdown(f"**Limitations:** {summary_json['limitations']}")
                            else:
                                st.markdown(gen_summary)
                        except (json.JSONDecodeError, TypeError):
                            st.markdown(gen_summary)

    except Exception as e:
        st.error(f"Could not render results: {e}")
